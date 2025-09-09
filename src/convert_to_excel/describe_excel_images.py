# 描述Excel中的图片

import os
import zipfile
from openpyxl import load_workbook
from PIL import Image
from utils.analyze_factory_image import analyze_factory_image
import logging

# 添加日志记录
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def process_excel(
    input_file: str = "input2.xlsx",
    target_column_name: str = "图片",
    description_column_name: str = "图片描述",
    output_img_dir: str = r"D:\CODE\excel\imgs",
    output_file: str = "output.xlsx",
    header_row: int = 1  # 可根据实际情况调整表头所在行
) -> None:
    """
    处理 Excel 文件，遍历每个sheet，在“图片”列中寻找图片，
    若找到则：
      1. 将图片保存到指定目录（默认 D:\CODE\excel\imgs）。
      2. 调用 analyze_factory_image 获取图片描述，并写入新列（默认“图片描述”）。
      
    参数:
      input_file: 输入 Excel 文件路径（默认为 input2.xlsx）
      target_column_name: 存放图片的列名（默认为“图片”）
      description_column_name: 图片描述存放的列名（默认为“图片描述”）
      output_img_dir: 图片保存目录（默认为 D:\CODE\excel\imgs）
      output_file: 输出的 Excel 文件路径（默认为 output.xlsx）
      header_row: 表头所在行（默认第 1 行）
    """
    # 确保图片保存目录存在
    os.makedirs(output_img_dir, exist_ok=True)

    try:
        wb = load_workbook(input_file)
    except zipfile.BadZipFile:
        logging.error("文件不是有效的 Excel 文件，请检查文件是否损坏或格式是否正确")
        return
    except FileNotFoundError:
        logging.error("文件未找到，请检查文件路径")
        return

    # 遍历每个 sheet
    for ws in wb.worksheets:
        
        logging.info(f"处理 sheet: {ws.title}")
        # 获取 header 行，注意空值的处理，并去除左右空白字符
        header = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[header_row]]
        if target_column_name not in header:
            logging.warning(f"Sheet {ws.title} 中未找到列 '{target_column_name}'，跳过该 sheet。")
            continue

        # openpyxl 中单元格列号从 1 开始；但 _images 中锚点属性 _from.col 是从 0 开始
        target_col_idx = header.index(target_column_name) + 1
        logging.info(f"目标列 '{target_column_name}' 对应 Excel 中的列号: {target_col_idx}")

        # 如果描述列不存在，则添加描述列（写入 header 行末尾）
        if description_column_name not in header:
            desc_col_idx = len(header) + 1
            ws.cell(row=header_row, column=desc_col_idx, value=description_column_name)
        else:
            desc_col_idx = header.index(description_column_name) + 1

        # 遍历当前 sheet 中所有图片
        for image in ws._images:
            try:
                # 获取图片的锚点信息：注意 row 和 col 均为 0 起始
                anchor_from = image.anchor._from
                img_row = anchor_from.row + 1  # 转换为 Excel 行号（从1开始）
                img_col = anchor_from.col + 1  # 转换为 Excel 列号（从1开始）
                logging.info(f"找到图片，锚点位于单元格 ({img_row}, {img_col})")
                
                # 调试信息：打印目标列与图片所在列
                if img_col == target_col_idx:
                    logging.info(f"图片在目标列 '{target_column_name}'（列号 {img_col}）")
                    # 构造图片文件名，确保唯一（以 sheet 名称和行号命名）
                    img_filename = f"{ws.title}_row{img_row}.png"
                    img_path = os.path.join(output_img_dir, img_filename)
                    img = Image.open(image.ref).convert("RGB")
                    img.save(img_path, format='PNG')
                    logging.info(f"图片已保存到 {img_path}")


                    try:
                        # 调用外部函数获取图片描述
                        description = analyze_factory_image(img_path)
                    except Exception as e:
                        logging.error(f"调用 analyze_factory_image 获取描述失败: {e}")
                        description = "描述获取失败"
                    
                    # 将描述写入对应行的描述列中
                    ws.cell(row=img_row, column=desc_col_idx, value=description)
                else:
                    logging.info(f"图片不在目标列: 图片所在列 {img_col} != 目标列 {target_col_idx}")
            except Exception as e:
                logging.error(f"处理图片时出错: {e}")
            

    # 保存并关闭处理后的 Excel 文件
    wb.save(output_file)
    wb.close()  # 显式关闭工作簿
    logging.info(f"处理完成，输出文件保存在 {output_file}")

if __name__ == "__main__":
    process_excel(
        r"output2\merged_manufacturers_final_concat.xlsx",
        target_column_name="图片1",
        description_column_name="图片描述",
        output_img_dir=r"output2\imgs",
        output_file=r"output2\output_img_desc.xlsx",
        header_row=1  
    )
